#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_catalog.py — Сборка каталога для promportal.su | ТД РУССтанкоСбыт

Источники (репозиторий Katalog-RSS):
  1. OFFERS_FILLED.csv           — основной (SKU, Title, Text, Description)
  2. MASTER_CATALOG_CLEANED.csv  — мастер-каталог
  3. CATALOG_OLEG_FINAL_100.xlsx — каталог Олега

Запуск:  pip install openpyxl pandas
         python build_catalog.py
Выход:   promportal_upload_ДАТА.xlsx
"""

import re, os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

# ── ПУТИ ──────────────────────────────────────────────────────────────────────
BASE = r"C:\GitHub-Repositories\Katalog-RSS"
SOURCES = {
    "offers_filled":  os.path.join(BASE, "OFFERS_FILLED.csv"),
    "master_cleaned": os.path.join(BASE, "MASTER_CATALOG_CLEANED.csv"),
    "catalog_xlsx":   os.path.join(BASE, "CATALOG_OLEG_FINAL_100.xlsx"),
}
OUTPUT_FILE = f"promportal_upload_{date.today().strftime('%Y%m%d')}.xlsx"

# ── МАШИНЫ ────────────────────────────────────────────────────────────────────
_MP = [
    r"1[МмMm][63]{2}[НнBбBFФ\d]*", r"1[НнNn][65]{2}[^\s,;]*",
    r"1[МмMm][65]{2}[^\s,;]*",     r"16[КкKk]20[^\s,;]*",
    r"16[КкKk]30[^\s,;]*",          r"16[КкKk]40[^\s,;]*",
    r"16[МмMm]30[^\s,;]*",          r"16[РрRr]25[^\s,;]*",
    r"1[КкKk]62[^\s,;]*",           r"РТ\d{3}[^\s,;]*",
    r"ДИП[35]\d{2}[^\s,;]*",       r"UBB\d{3}[^\s,;]*",
    r"1[АаAa]9[8]\d[^\s,;]*",      r"СА\d{3}[^\s,;]*",
]
MACHINE_RE = re.compile("|".join(_MP), re.IGNORECASE)

def extract_machines(text: str) -> str:
    found = MACHINE_RE.findall(text)
    unique = list(dict.fromkeys(re.sub(r"[,;.\s]+$", "", m) for m in found))
    return ", ".join(unique[:8])

# ── КАТЕГОРИИ ─────────────────────────────────────────────────────────────────
CATEGORY_MAP = [
    (["шпиндел"],                                   "Шпиндельная бабка"),
    (["фартук"],                                     "Фартук"),
    (["каретк", "суппорт", "салазк"],               "Суппорт и каретка"),
    (["задняя бабка", "задней бабк", "пиноль"],     "Задняя бабка"),
    (["шестерн", "зубчат", "колесо", "венец"],      "Шестерни и зубчатые колёса"),
    (["вал-шест", "вал шест", "вал-рейк", "вал "],  "Валы"),
    (["гайк", "маточн"],                             "Гайки"),
    (["винт"],                                       "Винты и ходовые винты"),
    (["люнет"],                                      "Люнеты"),
    (["патрон", "планшайб"],                         "Патроны"),
    (["резцедержат"],                                "Резцедержатели"),
    (["муфт"],                                       "Муфты"),
    (["кулачок", "кулачки", "кулачк"],              "Кулачки"),
    (["лимб"],                                       "Лимбы и нониусы"),
    (["клин"],                                       "Клинья"),
    (["насос", "помп", "смазк"],                     "Системы смазки"),
    (["ремонт", "восстановл"],                       "Ремонт и восстановление"),
    (["изготовл", "литьё", "литье", "отливк"],       "Производство на заказ"),
    (["швп", "шарико-винт"],                         "ШВП"),
    (["револьвер"],                                  "Револьверные головки"),
    (["захват"],                                     "Оснастка"),
    (["клапан", "гидро"],                            "Гидравлика"),
]
def detect_category(name: str) -> str:
    nl = name.lower()
    for kws, cat in CATEGORY_MAP:
        if any(k in nl for k in kws):
            return cat
    return "Запасные части для станков"

# ── ОЧИСТКА ───────────────────────────────────────────────────────────────────
SPAM = [
    "Запасные части , оснастка и комплектующие изделия для токарно",
    "Запасные части, оснастка и комплектующие изделия для токарно",
    "Запасные части для станков 1М63",
    "ВАЛЫ , ВАЛ - РЕЙКИ", "ВАЛЫ, ВАЛ - РЕЙКИ",
    "Тип предложения: предлагаю", "Товар на сайте компании",
]
def clean_text(text: str) -> str:
    if not text or str(text).strip() in ("nan", ""):
        return ""
    t = str(text)
    t = re.sub(r'<[^>]+>', ' ', t)
    t = t.replace("#nbsp;", " ").replace("&nbsp;", " ")
    t = re.sub(r'[ \t]+', ' ', t)
    for marker in SPAM:
        idx = t.find(marker)
        if idx > 60:
            t = t[:idx].strip(" ,.\n<>")
            break
    return t.strip()

def seo_title(name: str) -> str:
    n = clean_text(name)
    return n if len(n) <= 70 else n[:70].rsplit(" ", 1)[0].strip(" ,.")

def make_phrases(name: str, machines: str) -> str:
    p, cn = [], clean_text(name)
    if cn[:50]: p.append(cn[:50])
    words = cn.split()
    if words:
        b = f"купить {' '.join(words[:3])}"[:50]
        if b not in p: p.append(b)
    for m in (machines.split(", ")[:4] if machines else []):
        ph = f"запчасти {m}"[:50]
        if ph not in p: p.append(ph)
    for g in ["запчасти токарный станок","купить запчасти с доставкой","комплектующие токарного станка"]:
        if len(p) < 10 and g not in p: p.append(g)
    return "\n".join(p[:10])

# ── ПАРСЕРЫ ───────────────────────────────────────────────────────────────────
def parse_offers_csv(fp: str) -> list:
    if not os.path.exists(fp): print(f"  [SKIP] {fp}"); return []
    df = pd.read_csv(fp, dtype=str).fillna("")
    items = []
    for _, row in df.iterrows():
        name = row.get("Title", "").strip()
        if not name or name.lower() in ("nan","title"): continue
        raw_t = row.get("Text","").strip()
        raw_d = row.get("Description","").strip()
        desc  = clean_text(raw_t) if raw_t and raw_t != "nan" else clean_text(raw_d)
        img   = raw_d.strip() if raw_d.startswith("http") else ""
        items.append({"name":name,"desc":desc,"sku":row.get("SKU","").strip(),"image":img})
    print(f"  {os.path.basename(fp)}: {len(items)} позиций"); return items

def parse_master_csv(fp: str) -> list:
    if not os.path.exists(fp): print(f"  [SKIP] {fp}"); return []
    df = pd.read_csv(fp, dtype=str, low_memory=False).fillna("")
    cl = {c.lower(): c for c in df.columns}
    nc = next((cl[c] for c in cl if any(k in c for k in ["наимен","title","name","товар"])), None)
    dc = next((cl[c] for c in cl if any(k in c for k in ["описан","desc","text"])), None)
    sc = next((cl[c] for c in cl if any(k in c for k in ["артикул","sku","код"])), None)
    if not nc: print(f"  [WARN] {os.path.basename(fp)}: нет колонки имени"); return []
    items = []
    for _, row in df.iterrows():
        name = str(row[nc]).strip()
        if not name or name.lower() in ("nan","наименование","название"): continue
        items.append({"name":name,"desc":clean_text(str(row[dc])) if dc else "","sku":str(row[sc]).strip() if sc else "","image":""})
    print(f"  {os.path.basename(fp)}: {len(items)} позиций"); return items

def parse_xlsx(fp: str) -> list:
    if not os.path.exists(fp): print(f"  [SKIP] {fp}"); return []
    df = pd.read_excel(fp, dtype=str).fillna("")
    cl = {c.lower(): c for c in df.columns}
    nc = next((cl[c] for c in cl if any(k in c for k in ["наимен","title","name"])), None)
    dc = next((cl[c] for c in cl if any(k in c for k in ["описан","desc"])), None)
    sc = next((cl[c] for c in cl if any(k in c for k in ["артикул","sku","код"])), None)
    xc = next((cl[c] for c in cl if any(k in c for k in ["совмест","станок","модел"])), None)
    if not nc: print(f"  [WARN] {os.path.basename(fp)}: нет колонки имени"); return []
    items = []
    for _, row in df.iterrows():
        name = str(row[nc]).strip()
        if not name or name.lower() in ("nan","наименование"): continue
        items.append({"name":name,"desc":clean_text(str(row[dc])) if dc else "","sku":str(row[sc]).strip() if sc else "","compat":str(row[xc]).strip() if xc else "","image":""})
    print(f"  {os.path.basename(fp)}: {len(items)} позиций"); return items

# ── ДЕДУПЛИКАЦИЯ ──────────────────────────────────────────────────────────────
def normalize(s): return re.sub(r"\s+", " ", s.lower().strip(" .,;()"))
def deduplicate(sources):
    seen, result = set(), []
    for src in sources:
        for item in src:
            k = normalize(item["name"])
            if k in seen or not k: continue
            seen.add(k); result.append(item)
    print(f"  Всего на входе: {sum(len(s) for s in sources)}, уникальных: {len(result)}")
    return result

# ── XLSX ──────────────────────────────────────────────────────────────────────
HEADERS = ["Наименование","Описание","Единица измерения","Цена","Валюта",
           "Изображение","Раздел","Товарная группа","Наличие","Код товара",
           "Производитель","Поисковые фразы","Цена от","Цена со скидкой","Срок скидки",
           "Оптовая цена 1","Оптовый мин заказ 1","Оптовая цена 2","Оптовый мин заказ 2",
           "Оптовая цена 3","Оптовый мин заказ 3","Оптовая цена 4","Оптовый мин заказ 4",
           "Оптовая цена 5","Оптовый мин заказ 5"]
for i in range(1,11):
    HEADERS += [f"Название характеристики {i}",f"Значение характеристики {i}",f"Измерение характеристики {i}"]

def build_row(item):
    name  = item.get("name",""); desc = item.get("desc","")
    sku   = item.get("sku","");  compat = item.get("compat","")
    image = item.get("image","")
    title = seo_title(name)
    mach  = extract_machines(f"{name} {desc} {compat}")
    cat   = detect_category(name)
    phr   = make_phrases(name, mach)
    if not desc:
        desc = (f"{title}. Применяется на токарно-винторезных станках {mach}. " if mach else f"{title}. Запасная часть для токарного станка. ") + "Купить с доставкой по России. ТД РУССтанкоСбыт."
    row = [title,desc,"шт.","","руб.",image,"Запасные части для станков",cat,"под заказ",sku,"ТД РУССтанкоСбыт",phr,"","",""]
    row += [""]*10
    row += (["Совместимые станки",mach,""] if mach else ["","",""])
    row += (["Артикул производителя",sku,""] if sku else ["","",""])
    for _ in range(8): row += ["","",""]
    return row

def write_xlsx(items, out):
    wb = Workbook(); ws = wb.active; ws.title = "Лист1"
    ws.append(HEADERS)
    hf = PatternFill("solid",start_color="1F4E79"); hfont = Font(bold=True,color="FFFFFF",name="Arial",size=10)
    for cell in ws[1]:
        cell.fill=hf; cell.font=hfont; cell.alignment=Alignment(horizontal="center",wrap_text=True)
    rf = Font(name="Arial",size=10)
    for item in items:
        ws.append(build_row(item))
        for cell in ws[ws.max_row]:
            cell.font=rf; cell.alignment=Alignment(wrap_text=False,vertical="top")
    for col,w in {"A":65,"B":90,"G":35,"H":40,"L":45}.items():
        ws.column_dimensions[col].width=w
    ws2=wb.create_sheet("Товарные группы"); ws2.append(["Ид","Наименование","Ид родителя"])
    wb.save(out); print(f"\n✅ {out} ({len(items)} строк)")

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("="*60+"\nСборка каталога promportal.su — ТД РУССтанкоСбыт\n"+"="*60)
    print("\n[1/3] Читаю источники...")
    src = [parse_offers_csv(SOURCES["offers_filled"]),
           parse_master_csv(SOURCES["master_cleaned"]),
           parse_xlsx(SOURCES["catalog_xlsx"])]
    print("\n[2/3] Дедупликация...")
    merged = deduplicate(src)
    print(f"\n[3/3] Генерирую {OUTPUT_FILE}...")
    write_xlsx(merged, OUTPUT_FILE)
    print("\nГотово! Загрузи через Импорт товаров на promportal.su")

if __name__ == "__main__":
    main()
