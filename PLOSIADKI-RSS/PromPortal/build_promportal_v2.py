#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_promportal.py — Финальный сборщик каталога для promportal.su
ТД РУССтанкоСбыт

Источники:
  1. DIRECTUS_TKP_549_FULL.csv  — Tilda-экспорт (приоритет: есть фото, цена, описание)
  2. CATALOG_OLEG_FINAL_100.xlsx — каталог Олега
  3. NOMENCLATURA RSS-2026.md   — 2828 позиций (имена)
  4. RZN-список офферов.md      — офферы РЗН
  5. Запасные части...1М63Н.md  — запчасти по моделям

Исключения:
  СПИСОК ОФФЕРОВ.md             — уже загружено (не дублировать)

Запуск:
  pip install openpyxl pandas
  python build_promportal.py

Результат:
  PROMPORTAL_READY_ДАТА.xlsx    — готов к импорту на promportal.su
"""

import re
import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

# ══════════════════════════════════════════════════════════════════
# КОНФИГ — меняй только здесь
# ══════════════════════════════════════════════════════════════════
BASE = r"C:\GitHub-Repositories\Katalog-RSS"

SOURCES = {
    # Приоритет 1 — Tilda-экспорт: SKU, цена, фото, описание, совместимость
    "tilda_csv": r"C:\GitHub-Repositories\Katalog-RSS\Directus-RSS\Товары\Каталоги CSV\DIRECTUS_TKP_549_FULL.csv",

    # Приоритет 2 — 2828 уникальных позиций
    "nomenclatura_md": r"C:\GitHub-Repositories\Katalog-RSS\NOMENCLATURA RSS-2026.md",

    # Приоритет 3 — офферы РЗН
    "rzn_offers_md": r"C:\GitHub-Repositories\Katalog-RSS\PLOSIADKI-RSS\PromPortal\RZN-список офферов.md",

    # Приоритет 4 — запчасти по моделям
    "parts_1m63_md": r"C:\GitHub-Repositories\Katalog-RSS\RZN-сайт\Запасные части для токарно-винторезных станков мод. 1М63Н(1М63),16К40 (1А64),1Н65(1М65),РТ117,РТ817.md",
}

EXCLUDE_FILE = r"C:\GitHub-Repositories\Katalog-RSS\PLOSIADKI-RSS\PromPortal\СПИСОК ОФФЕРОВ.md"
OUTPUT_FILE  = r"C:\GitHub-Repositories\Katalog-RSS\PLOSIADKI-RSS\PromPortal\СПИСОК ОФФЕРОВ+ОПИСАНИЕ.xlsx"

# ══════════════════════════════════════════════════════════════════
# ТОЧНЫЕ ЗАГОЛОВКИ ШАБЛОНА promportal.su (55 колонок, порядок строгий)
# ══════════════════════════════════════════════════════════════════
PROMPORTAL_HEADERS = [
    "Наименование", "Описание", "Единица измерения", "Цена", "Валюта",
    "Изображение", "Раздел", "Товарная группа", "Наличие", "Код товара",
    "Производитель", "Поисковые фразы", "Цена от", "Цена со скидкой", "Срок скидки",
    "Оптовая цена 1",    "Оптовый мин заказ 1",
    "Оптовая цена 2",    "Оптовый мин заказ 2",
    "Оптовая цена 3",    "Оптовый мин заказ 3",
    "Оптовая цена 4",    "Оптовый мин заказ 4",
    "Оптовая цена 5",    "Оптовый мин заказ 5",
    "Название характеристики 1",  "Значение характеристики 1",  "Измерение характеристики 1",
    "Название характеристики 2",  "Значение характеристики 2",  "Измерение характеристики 2",
    "Название характеристики 3",  "Значение характеристики 3",  "Измерение характеристики 3",
    "Название характеристики 4",  "Значение характеристики 4",  "Измерение характеристики 4",
    "Название характеристики 5",  "Значение характеристики 5",  "Измерение характеристики 5",
    "Название характеристики 6",  "Значение характеристики 6",  "Измерение характеристики 6",
    "Название характеристики 7",  "Значение характеристики 7",  "Измерение характеристики 7",
    "Название характеристики 8",  "Значение характеристики 8",  "Измерение характеристики 8",
    "Название характеристики 9",  "Значение характеристики 9",  "Измерение характеристики 9",
    "Название характеристики 10", "Значение характеристики 10", "Измерение характеристики 10",
]
assert len(PROMPORTAL_HEADERS) == 55, f"Ошибка: {len(PROMPORTAL_HEADERS)} колонок вместо 55"

# ══════════════════════════════════════════════════════════════════
# REGEXP: извлечение моделей станков
# ══════════════════════════════════════════════════════════════════
_MACH_PATS = [
    r"1[МмMm][63]{2}[НнБбBFФ\d]*",
    r"1[НнNn][65]{2}[^\s,;]*",
    r"1[МмMm][65]{2}[^\s,;]*",
    r"16[КкKk]20[^\s,;]*",
    r"16[КкKk]30[^\s,;]*",
    r"16[КкKk]40[^\s,;]*",
    r"16[АаAa]20[^\s,;]*",
    r"16[МмMm]30[^\s,;]*",
    r"16[РрRr]25[^\s,;]*",
    r"1[КкKk]62[^\s,;]*",
    r"2[МмMm]\d{2,3}[^\s,;]*",
    r"РТ\d{3}[^\s,;]*",
    r"ДИП[345]\d{2}[^\s,;]*",
    r"УГ\d{4}[^\s,;]*",
    r"СА\d{3}[^\s,;]*",
    r"1[АаAa]\d{3}[^\s,;]*",
    r"1[НнNn]\d{3}[^\s,;]*",
    r"FSS\s*\d{3}[^\s,;]*",
    r"UBB\s*\d{3}[^\s,;]*",
    r"6[РрTтNнМм]\d{2}[^\s,;]*",
    r"2825П[^\s,;]*",
]
MACH_RE = re.compile("|".join(_MACH_PATS), re.IGNORECASE)

def extract_machines(text: str) -> str:
    if not text:
        return ""
    hits = MACH_RE.findall(text)
    clean = [re.sub(r"[,;.\s]+$", "", m) for m in hits]
    unique = list(dict.fromkeys(c for c in clean if len(c) >= 3))
    return ", ".join(unique[:10])

# ══════════════════════════════════════════════════════════════════
# КАТЕГОРИИ
# ══════════════════════════════════════════════════════════════════
_CAT = [
    (["шпиндел"],                                "Узлы токарных станков",   "Шпиндельные бабки"),
    (["фартук"],                                 "Узлы токарных станков",   "Фартуки"),
    (["каретк", "суппорт", "салазк", "ползуш"], "Узлы токарных станков",   "Каретки и суппорты"),
    (["задняя бабка", "задней бабк", "пиноль"], "Узлы токарных станков",   "Задние бабки"),
    (["коробка подач", "коробка скоростей"],     "Узлы токарных станков",   "Коробки передач"),
    (["ходовой винт", "ходовой вал"],            "Узлы токарных станков",   "Винты, ШВП, валы"),
    (["швп", "шарико-винт"],                     "Узлы токарных станков",   "Винты, ШВП, валы"),
    (["шестерн", "зубчат", "колесо", "венец"],   "Комплектующие к узлам",   "Шестерни и зубчатые колёса"),
    (["вал-шест", "вал шест"],                   "Комплектующие к узлам",   "Шестерни и зубчатые колёса"),
    (["вал "],                                   "Комплектующие к узлам",   "Валы"),
    (["гайк", "маточн"],                         "Комплектующие к узлам",   "Гайки"),
    (["винт"],                                   "Узлы токарных станков",   "Винты, ШВП, валы"),
    (["люнет"],                                  "Оснастка токарная",       "Люнеты"),
    (["патрон", "планшайб"],                     "Оснастка токарная",       "Патроны токарные"),
    (["резцедержат"],                            "Оснастка токарная",       "Резцедержатели"),
    (["муфт"],                                   "Комплектующие к узлам",   "Муфты"),
    (["кулачок", "кулачки", "кулачк"],           "Оснастка токарная",       "Кулачки"),
    (["лимб"],                                   "Комплектующие к узлам",   "Лимбы и нониусы"),
    (["клин"],                                   "Комплектующие к узлам",   "Клинья"),
    (["насос", "смазк"],                         "Комплектующие к узлам",   "Системы смазки"),
    (["ремонт", "восстановл"],                   "Услуги",                  "Ремонт и восстановление"),
    (["изготовл"],                               "Услуги",                  "Производство на заказ"),
    (["революьвер", "револьвер"],                "Узлы токарных станков",   "Револьверные головки"),
    (["блок зуб", "блок-шест", "блок шест"],     "Комплектующие к узлам",   "Шестерни и зубчатые колёса"),
    (["барабан"],                                "Комплектующие к узлам",   "Прочие комплектующие"),
]
def detect_category(name: str) -> tuple:
    nl = name.lower()
    for kws, section, group in _CAT:
        if any(k in nl for k in kws):
            return section, group
    return "Запасные части для станков", "Комплектующие токарных станков"

# ══════════════════════════════════════════════════════════════════
# ОЧИСТКА ТЕКСТА
# ══════════════════════════════════════════════════════════════════
_SPAM_MARKERS = [
    "Смотрите также:", "Запасные части , оснастка",
    "Запасные части, оснастка", "Тип предложения:",
    "Товар на сайте", "ВАЛЫ , ВАЛ - РЕЙКИ",
]
def clean_text(raw) -> str:
    if not raw or str(raw).strip() in ("", "nan"):
        return ""
    t = str(raw)
    t = re.sub(r"<[^>]+>", " ", t)
    t = t.replace("&nbsp;", " ").replace("#nbsp;", " ")
    t = re.sub(r"[ \t]+", " ", t)
    for marker in _SPAM_MARKERS:
        idx = t.find(marker)
        if idx > 40:
            t = t[:idx].strip(" ,.\n")
            break
    return t.strip()

# ══════════════════════════════════════════════════════════════════
# ПОИСКОВЫЕ ФРАЗЫ (одна колонка, разделитель \n)
# ══════════════════════════════════════════════════════════════════
def make_phrases(name: str, machines: str, sku: str) -> str:
    p = []
    cn = name.strip()
    if cn[:50]: p.append(cn[:50])
    words = cn.split()
    buy = f"купить {' '.join(words[:3])}"[:50]
    if buy not in p: p.append(buy)
    for m in (machines.split(", ")[:3] if machines else []):
        ph = f"запчасти {m}"[:50]
        if ph not in p: p.append(ph)
    if sku and len(sku) > 3:
        ph = f"{sku[:40]} купить"[:50]
        if ph not in p: p.append(ph)
    for g in ["запчасти токарный станок", "купить запчасти с доставкой",
              "комплектующие токарного станка"]:
        if len(p) < 10 and g not in p:
            p.append(g)
    return "\n".join(p[:10])

# ══════════════════════════════════════════════════════════════════
# ГЕНЕРАЦИЯ ОПИСАНИЯ (если нет готового)
# ══════════════════════════════════════════════════════════════════
def auto_desc(name: str, machines: str, sku: str) -> str:
    mach_str = f" {machines}." if machines else ""
    sku_str  = f" Артикул: {sku}." if sku and sku != "nan" else ""
    return (f"{name} — запасная часть для токарного станка.{mach_str}"
            f"{sku_str} Купить с доставкой по России. "
            f"Подбор по каталогу и чертежам. ТД РУССтанкоСбыт.")

# ══════════════════════════════════════════════════════════════════
# ПАРСЕРЫ ИСТОЧНИКОВ
# ══════════════════════════════════════════════════════════════════

def parse_tilda_csv(fp: str) -> list:
    """
    Tilda CSV: Tilda UID | Brand | SKU | ... | Title | Description | Text |
               Photo | Price | ... | Tabs:1 | meta_title | meta_description
    """
    if not os.path.exists(fp):
        print(f"  [SKIP] {fp}"); return []
    df = pd.read_csv(fp, dtype=str, low_memory=False).fillna("")
    cols = {c.strip(): c for c in df.columns}

    def col(*names):
        for n in names:
            if n in cols: return cols[n]
        # partial match
        for n in names:
            for k in cols:
                if n.lower() in k.lower(): return cols[k]
        return None

    c_sku   = col("SKU", "Артикул")
    c_title = col("Title", "Наименование", "Название")
    c_text  = col("Text", "Описание подробное")
    c_desc  = col("Description", "Описание краткое")
    c_photo = col("Photo", "Фото")
    c_price = col("Price", "Цена")
    c_cat   = col("Category", "Категория")
    c_tabs  = col("Tabs:1")   # тут живёт "Совместимость: ..."

    items = []
    for _, row in df.iterrows():
        title = str(row[c_title]).strip() if c_title else ""
        if not title or title.lower() in ("nan", "title", "наименование"):
            continue

        text  = clean_text(row[c_text])  if c_text  else ""
        desc  = clean_text(row[c_desc])  if c_desc  else ""
        body  = text if text else desc

        photo = str(row[c_photo]).strip() if c_photo else ""
        if photo.lower() in ("nan", ""): photo = ""

        price = str(row[c_price]).strip() if c_price else ""
        if price in ("nan", "0", ""): price = ""

        sku   = str(row[c_sku]).strip() if c_sku else ""
        if sku == "nan": sku = ""

        # Совместимые модели из Tabs:1 — формат "Характеристики\nСовместимость: 1М63;16К20;..."
        compat = ""
        if c_tabs:
            tabs_raw = str(row[c_tabs])
            m = re.search(r"Совместимость:\s*([^\n]+)", tabs_raw)
            if m:
                compat = m.group(1).replace(";", ", ").strip()

        # Раздел / Товарная группа из Category "Раздел>>>Группа"
        cat_raw = str(row[c_cat]).strip() if c_cat else ""
        if ">>>" in cat_raw:
            parts = cat_raw.split(">>>")
            cat_section = parts[0].strip()
            cat_group   = parts[-1].strip()
        else:
            cat_section, cat_group = detect_category(title)

        items.append({
            "name": title, "desc": body, "sku": sku,
            "photo": photo, "price": price,
            "compat": compat,
            "section": cat_section, "group": cat_group,
        })
    print(f"  DIRECTUS_TKP_549_FULL.csv: {len(items)} позиций")
    return items


def parse_oleg_xlsx(fp: str) -> list:
    """
    Колонки: ID, Title, SKU, Type, Brand, Category_Path, Price, Currency,
             Compatible_SKUs, Short_Description, Full_Description, ..., Main_Photo, Status
    """
    if not os.path.exists(fp):
        print(f"  [SKIP] {fp}"); return []
    df = pd.read_excel(fp, dtype=str).fillna("")
    cl = {c.strip().lower(): c for c in df.columns}

    def col(*names):
        for n in names:
            if n.lower() in cl: return cl[n.lower()]
        for n in names:
            for k in cl:
                if n.lower() in k: return cl[k]
        return None

    c_title  = col("Title", "Наименование")
    c_sku    = col("SKU", "Артикул", "Код товара")
    c_price  = col("Price", "Цена")
    c_compat = col("Compatible_SKUs", "Совместимость", "Совместимые")
    c_desc   = col("Full_Description", "Описание", "Description")
    c_photo  = col("Main_Photo", "Фото", "Photo")
    c_cat    = col("Category_Path", "Категория")

    if not c_title:
        print(f"  [WARN] CATALOG_OLEG_FINAL_100.xlsx: не найдена колонка Title"); return []

    items = []
    for _, row in df.iterrows():
        title = str(row[c_title]).strip()
        if not title or title.lower() in ("nan", "title", "название товара"):
            continue
        sku   = str(row[c_sku]).strip()    if c_sku    else ""
        price = str(row[c_price]).strip()  if c_price  else ""
        compat= str(row[c_compat]).strip() if c_compat else ""
        body  = clean_text(row[c_desc])    if c_desc   else ""
        photo = str(row[c_photo]).strip()  if c_photo  else ""

        for v in ("nan", "0"): price = "" if price == v else price
        for v in ("nan",):
            sku   = "" if sku   == v else sku
            compat= "" if compat == v else compat
            photo = "" if photo == v else photo

        # Compat может быть "1М63;16К20;1Н65" → нормализуем
        compat = compat.replace(";", ", ")

        cat_raw = str(row[c_cat]).strip() if c_cat else ""
        if ">>>" in cat_raw:
            parts = cat_raw.split(">>>")
            cat_section = parts[0].strip()
            cat_group   = parts[-1].strip()
        else:
            cat_section, cat_group = detect_category(title)

        items.append({
            "name": title, "desc": body, "sku": sku,
            "photo": photo, "price": price,
            "compat": compat,
            "section": cat_section, "group": cat_group,
        })
    print(f"  CATALOG_OLEG_FINAL_100.xlsx: {len(items)} позиций")
    return items


_JUNK_LINE = re.compile(
    r"^(\d[\d\s\u2009]*[₽/]|"      # цены
    r"\d+\s*(шт|кг|п\.м|комплект)|"  # кол-во
    r"[-–—]{3,}|"                   # разделители
    r"#{1,6}\s|"                    # заголовки md
    r"\|)"                          # таблицы
    , re.IGNORECASE
)
def parse_md(fp: str) -> list:
    if not os.path.exists(fp):
        print(f"  [SKIP] {fp}"); return []
    items = []
    cur_title, cur_desc = None, []
    with open(fp, encoding="utf-8", errors="ignore") as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue
            if _JUNK_LINE.match(line):
                continue
            # Длинная строка (≥120) = продолжение описания
            if len(line) >= 120 and cur_title:
                cur_desc.append(clean_text(line)); continue
            # Строка начинается с буквы — новый товар
            if re.match(r"^[А-ЯЁа-яёA-Za-z(«]", line) and len(line) > 5:
                if cur_title:
                    items.append({"name": cur_title,
                                  "desc": " ".join(cur_desc),
                                  "sku": "", "photo": "", "price": "",
                                  "compat": "", "section": "", "group": ""})
                cur_title = line; cur_desc = []
    if cur_title:
        items.append({"name": cur_title, "desc": " ".join(cur_desc),
                      "sku": "", "photo": "", "price": "",
                      "compat": "", "section": "", "group": ""})
    print(f"  {os.path.basename(fp)}: {len(items)} позиций")
    return items


def load_exclude(fp: str) -> set:
    if not os.path.exists(fp):
        print(f"  [WARN] файл исключений не найден: {fp}"); return set()
    items = parse_md(fp)
    ex = {_norm(i["name"]) for i in items}
    print(f"  Исключений (уже на портале): {len(ex)}")
    return ex


# ══════════════════════════════════════════════════════════════════
# ДЕДУПЛИКАЦИЯ
# ══════════════════════════════════════════════════════════════════
def _norm(s: str) -> str:
    s = s.lower().strip(" .,;()")
    s = re.sub(r"\s+", " ", s)
    # убираем типичные стоп-слова чтобы "вал 1м63" == "вал для станка 1м63"
    s = re.sub(r"\bдля\b|\bстанк[аов]\b|\bмод\b\.?|\bмодел[ьи]\b", "", s)
    return re.sub(r"\s+", " ", s).strip()

def deduplicate(sources: list, exclude: set) -> list:
    seen   = set(exclude)
    result = []
    total  = sum(len(s) for s in sources)
    for src in sources:
        for item in src:
            k = _norm(item["name"])
            if not k or k in seen:
                continue
            seen.add(k)
            result.append(item)
    print(f"  Всего на входе: {total}")
    print(f"  Уникальных новых: {len(result)}")
    print(f"  Пропущено (дубли + исключения): {total - len(result)}")
    return result


# ══════════════════════════════════════════════════════════════════
# СБОРКА СТРОКИ ДЛЯ PROMPORTAL (55 значений)
# ══════════════════════════════════════════════════════════════════
def build_row(item: dict) -> list:
    name   = item["name"].strip()
    desc   = item["desc"]
    sku    = item["sku"]
    photo  = item["photo"]
    price  = item["price"]
    compat = item["compat"]

    # Машины: сначала из поля compat, потом regex по всему тексту
    machines = compat if compat else extract_machines(f"{name} {desc}")

    # Категория: приоритет у сохранённой из источника
    section = item.get("section") or ""
    group   = item.get("group")   or ""
    if not section or not group:
        section, group = detect_category(name)

    # Описание
    if not desc:
        desc = auto_desc(name, machines, sku)

    phrases = make_phrases(name, machines, sku)

    # Базовые 15 полей
    row = [
        name,           # 1  Наименование
        desc,           # 2  Описание
        "шт.",          # 3  Единица измерения
        price,          # 4  Цена
        "руб.",         # 5  Валюта
        photo,          # 6  Изображение
        section,        # 7  Раздел
        group,          # 8  Товарная группа
        "под заказ",    # 9  Наличие
        sku,            # 10 Код товара
        "ТД РУССтанкоСбыт",  # 11 Производитель
        phrases,        # 12 Поисковые фразы
        "",             # 13 Цена от
        "",             # 14 Цена со скидкой
        "",             # 15 Срок скидки
    ]
    # 16-25: 5 оптовых цен × 2
    row += [""] * 10

    # 26-55: 10 характеристик × 3
    chars = []
    if machines:
        chars.append(("Совместимые станки", machines, ""))
    if sku and sku not in ("nan", ""):
        chars.append(("Артикул производителя", sku, ""))
    # заполнить до 10
    while len(chars) < 10:
        chars.append(("", "", ""))
    for name_c, val_c, unit_c in chars[:10]:
        row += [name_c, val_c, unit_c]

    assert len(row) == 55, f"Строка: {len(row)} полей вместо 55"
    return row


# ══════════════════════════════════════════════════════════════════
# ЗАПИСЬ XLSX
# ══════════════════════════════════════════════════════════════════
def write_xlsx(items: list, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    # Заголовок
    ws.append(PROMPORTAL_HEADERS)
    hfill = PatternFill("solid", start_color="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Данные
    rfont = Font(name="Arial", size=10)
    for item in items:
        ws.append(build_row(item))
        for cell in ws[ws.max_row]:
            cell.font = rfont
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    # Ширина ключевых колонок
    widths = {"A": 65, "B": 90, "F": 50, "G": 38, "H": 42, "J": 30, "L": 45}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # Лист "Товарные группы" — пустой, нужен для импорта
    ws2 = wb.create_sheet("Товарные группы")
    ws2.append(["Ид", "Наименование", "Ид родителя"])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"\n  Файл: {out_path}")
    print(f"  Строк: {len(items)}")


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════
def main():
    print("=" * 65)
    print("  ТД РУССтанкоСбыт — Сборка каталога promportal.su")
    print("=" * 65)

    print("\n[1/4] Загружаю список исключений...")
    exclude = load_exclude(EXCLUDE_FILE)

    print("\n[2/4] Читаю источники (приоритет: Tilda > Олег > MD)...")
    sources = [
        parse_tilda_csv(SOURCES["tilda_csv"]),        # богатый источник
        parse_md(SOURCES["nomenclatura_md"]),          # 2828 имён
        parse_md(SOURCES["rzn_offers_md"]),            # офферы РЗН
        parse_md(SOURCES["parts_1m63_md"]),            # запчасти 1М63Н
    ]

    print("\n[3/4] Дедупликация...")
    merged = deduplicate(sources, exclude)

    print("\n[4/4] Генерирую XLSX...")
    write_xlsx(merged, OUTPUT_FILE)

    # Статистика
    df_stat = pd.DataFrame([{"s": i.get("section","?"), "g": i.get("group","?")}
                             for i in merged])
    print("\n📊 Разделы:")
    print(df_stat["s"].value_counts().to_string())
    print("\n📊 Топ-15 товарных групп:")
    print(df_stat["g"].value_counts().head(15).to_string())
    print("\n✅ Готово! Загрузи через Импорт товаров на promportal.su")


if __name__ == "__main__":
    main()
