#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_promportal_v3.py — Сборщик каталога для promportal.su
ТД РУССтанкоСбыт

Источники обогащения (приоритет):
  1. Directus-офферы сайта/*.md  — готовые карточки: описание, meta, тех.хар-ки
  2. DIRECTUS_TKP_549_FULL.csv   — фото, цены, совместимость
  3. Kinematika-Chertegi/*.xls   — число зубьев, модуль, материал
  4. Узлы, детали и цены.md      — ценовые ориентиры и ТОП-номенклатура
  5. MD-файлы номенклатуры       — имена без описаний

Исключения:  СПИСОК ОФФЕРОВ.md  — уже загружено
Результат:   СПИСОК ОФФЕРОВ+ОПИСАНИЕ.xlsx (55 колонок promportal + meta_title + meta_description)
"""

import re, os, glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date
from pathlib import Path

# ══════════════════════════════════════════════════════════════════
# ПУТИ
# ══════════════════════════════════════════════════════════════════
BASE   = Path(r"C:\GitHub-Repositories\Katalog-RSS")
DOFFERS_DIR  = BASE / "Directus-RSS" / "Товары" / "Directus-офферы сайта"
TILDA_CSV    = BASE / "Directus-RSS" / "Товары" / "Каталоги CSV" / "DIRECTUS_TKP_549_FULL.csv"
KINEMATICS_DIR = BASE / "Kinematika-Chertegi"
UZLY_MD      = BASE / "Kinematika-Chertegi" / "Узлы, детали и цены.md"
NOMENCLATURA = BASE / "NOMENCLATURA RSS-2026.md"
RZN_OFFERS   = BASE / "PLOSIADKI-RSS" / "PromPortal" / "RZN-список офферов.md"
PARTS_1M63   = BASE / "RZN-сайт" / "Запасные части для токарно-винторезных станков мод. 1М63Н(1М63),16К40 (1А64),1Н65(1М65),РТ117,РТ817.md"
EXCLUDE_FILE = BASE / "PLOSIADKI-RSS" / "PromPortal" / "СПИСОК ОФФЕРОВ.md"
OUTPUT_FILE  = BASE / "PLOSIADKI-RSS" / "PromPortal" / "СПИСОК ОФФЕРОВ+ОПИСАНИЕ.xlsx"

# ══════════════════════════════════════════════════════════════════
# ЗАГОЛОВКИ promportal.su (55) + наши meta (2) = 57
# ══════════════════════════════════════════════════════════════════
PROMPORTAL_HEADERS = [
    "Наименование","Описание","Единица измерения","Цена","Валюта",
    "Изображение","Раздел","Товарная группа","Наличие","Код товара",
    "Производитель","Поисковые фразы","Цена от","Цена со скидкой","Срок скидки",
    "Оптовая цена 1","Оптовый мин заказ 1","Оптовая цена 2","Оптовый мин заказ 2",
    "Оптовая цена 3","Оптовый мин заказ 3","Оптовая цена 4","Оптовый мин заказ 4",
    "Оптовая цена 5","Оптовый мин заказ 5",
    "Название характеристики 1","Значение характеристики 1","Измерение характеристики 1",
    "Название характеристики 2","Значение характеристики 2","Измерение характеристики 2",
    "Название характеристики 3","Значение характеристики 3","Измерение характеристики 3",
    "Название характеристики 4","Значение характеристики 4","Измерение характеристики 4",
    "Название характеристики 5","Значение характеристики 5","Измерение характеристики 5",
    "Название характеристики 6","Значение характеристики 6","Измерение характеристики 6",
    "Название характеристики 7","Значение характеристики 7","Измерение характеристики 7",
    "Название характеристики 8","Значение характеристики 8","Измерение характеристики 8",
    "Название характеристики 9","Значение характеристики 9","Измерение характеристики 9",
    "Название характеристики 10","Значение характеристики 10","Измерение характеристики 10",
    # Доп. колонки для SEO (не в шаблоне promportal — для справки)
    "META_TITLE","META_DESCRIPTION",
]
assert len(PROMPORTAL_HEADERS) == 57

# ══════════════════════════════════════════════════════════════════
# 14 КАТЕГОРИЙ (по реальным группам russtanko-rzn.ru)
# ══════════════════════════════════════════════════════════════════
CATEGORIES = [
    (["вал-рейк","вал рейк","рейка"],                          "Узлы токарных станков","Валы, вал-рейки, вал-шестерни"),
    (["вал-шест","вал шест","вал-колесо"],                     "Узлы токарных станков","Валы, вал-рейки, вал-шестерни"),
    (["вал ","валик ","вал\t"],                                "Узлы токарных станков","Валы, вал-рейки, вал-шестерни"),
    (["винт ход","ходовой винт","винт поперечн","кареточный винт","винт с гайк"],
                                                               "Узлы токарных станков","Винты ходовые и поперечных подач"),
    (["швп","шарико-винт","шарикова"],                         "Узлы токарных станков","Шарико-винтовые пары (ШВП)"),
    (["гайка маточн","маточная гайка","полугайк"],             "Узлы токарных станков","Гайки маточные"),
    (["диск фрикц","фрикционный диск"],                        "Узлы токарных станков","Диски фрикционные"),
    (["колесо зубч","зубчатое колесо","колесо конич","коническая шестерн","шестерн","зубчат","венец","червячн"],
                                                               "Узлы токарных станков","Колёса зубчатые, конические, червячные пары"),
    (["кулачок","кулачки","кулачк"],                           "Оснастка токарная",   "Кулачки к патронам"),
    (["насос масл","насос шестер","маслонасос"],                "Узлы токарных станков","Насосы масляные"),
    (["муфта обгон","обгонная муфта"],                         "Узлы токарных станков","Обгонные муфты"),
    (["муфта фрикц","фрикционная муфта","муфта тормоз","электромагн муфт"],
                                                               "Узлы токарных станков","Фрикционные муфты"),
    (["муфт"],                                                 "Узлы токарных станков","Муфты"),
    (["револьвер"],                                            "Узлы токарных станков","Револьверные головки"),
    (["резцедержат"],                                          "Оснастка токарная",   "Резцедержатели"),
    (["шкив"],                                                 "Узлы токарных станков","Шкивы главного привода"),
    (["шпиндел"],                                              "Узлы токарных станков","Шпиндельные бабки"),
    (["фартук"],                                               "Узлы токарных станков","Фартуки"),
    (["каретк","суппорт","салазк","ползуш"],                   "Узлы токарных станков","Каретки и суппорты"),
    (["задняя бабка","задней бабк","пиноль","бабка задн"],     "Узлы токарных станков","Задние бабки"),
    (["коробка подач","коробка скорост"],                      "Узлы токарных станков","Коробки передач"),
    (["патрон","планшайб"],                                    "Оснастка токарная",   "Патроны токарные"),
    (["люнет"],                                                "Оснастка токарная",   "Люнеты"),
    (["втулка перех","втулка конусн"],                         "Оснастка токарная",   "Втулки переходные"),
    (["центр токарн","центры и перех"],                        "Оснастка токарная",   "Центры и втулки"),
    (["гайк","полугайк"],                                      "Узлы токарных станков","Гайки"),
    (["винт"],                                                 "Узлы токарных станков","Винты и крепёж"),
    (["ремонт","восстановл","капитальн"],                      "Услуги",              "Ремонт и восстановление"),
    (["изготовл","производств"],                               "Услуги",              "Производство на заказ"),
    (["лимб"],                                                 "Узлы токарных станков","Лимбы и нониусы"),
    (["клин "],                                                "Узлы токарных станков","Клинья"),
    (["ось ","осевой"],                                        "Узлы токарных станков","Оси"),
    (["рейка "],                                               "Узлы токарных станков","Рейки"),
    (["насос"],                                                "Узлы токарных станков","Насосы"),
    (["подшипник"],                                            "Узлы токарных станков","Подшипники"),
    (["блок зуб","блок-шест","блок шест"],                     "Узлы токарных станков","Колёса зубчатые, конические, червячные пары"),
]
def detect_category(name: str) -> tuple:
    nl = name.lower()
    for kws, section, group in CATEGORIES:
        if any(k in nl for k in kws):
            return section, group
    return "Запасные части для станков", "Комплектующие токарных станков"

# ══════════════════════════════════════════════════════════════════
# REGEXP
# ══════════════════════════════════════════════════════════════════
MACH_RE = re.compile(
    r"1[МмMm][63]{2}[НнБбBFФ\d]*|1[НнNn][65]{2}[^\s,;]*|1[МмMm][65]{2}[^\s,;]*"
    r"|16[КкKk]20[^\s,;]*|16[КкKk]30[^\s,;]*|16[КкKk]40[^\s,;]*"
    r"|16[АаAa]20[^\s,;]*|16[МмMm]30[^\s,;]*|16[РрRr]25[^\s,;]*"
    r"|1[КкKk]62[^\s,;]*|РТ\d{3}[^\s,;]*|ДИП[345]\d{2}[^\s,;]*"
    r"|СА\d{3}[^\s,;]*|1[АаAa]\d{3}[^\s,;]*|1[НнNn]\d{3}[^\s,;]*"
    r"|FSS\s*\d{3}[^\s,;]*|UBB\s*\d{3}[^\s,;]*|6[РрTтNнМм]\d{2}[^\s,;]*"
    r"|2[МмMm]\d{2,3}[^\s,;]*|2825П[^\s,;]*|У05\.\d+[^\s,;]*",
    re.IGNORECASE
)
SKU_RE = re.compile(
    r"\b([А-ЯЁA-Z0-9][А-ЯЁA-Z0-9]*[\.\-][0-9]{2,}[\.\-][0-9]{2,}[^\s,;]*)",
    re.IGNORECASE
)
TEETH_RE  = re.compile(r"[Чч]исло\s+зубьев[:\s]+(\d+)")
MODULE_RE = re.compile(r"[Мм]одуль[:\s]+([\d,\.]+)")
MATERIAL_RE = re.compile(r"[Мм]атериал[:\s]+([^\n•\*]{5,60})")
HEAT_RE   = re.compile(r"[Тт]ермообработк[а-я]+[:\s]+([^\n•\*]{5,80})")
HRC_RE    = re.compile(r"HRC\s*([\d\-–]+)")
PRICE_RE  = re.compile(r"[\d\s\u2009]+[₽руб]")

SPAM = ["Смотрите также:", "Запасные части , оснастка",
        "Запасные части, оснастка", "Тип предложения:"]

def clean(raw) -> str:
    if not raw or str(raw).strip() in ("", "nan"): return ""
    t = re.sub(r"<[^>]+>", " ", str(raw))
    t = t.replace("&nbsp;"," ").replace("#nbsp;"," ")
    t = re.sub(r"[ \t]+"," ", t)
    for m in SPAM:
        i = t.find(m)
        if i > 40: t = t[:i].strip(" ,.\n"); break
    return t.strip()

def extract_machines(text: str) -> str:
    if not text: return ""
    hits = MACH_RE.findall(text)
    u = list(dict.fromkeys(re.sub(r"[,;.\s]+$","",h) for h in hits if len(h)>=3))
    return ", ".join(u[:10])

def extract_sku(text: str) -> str:
    m = SKU_RE.search(text)
    return m.group(1).strip(".,;") if m else ""

def norm(s: str) -> str:
    s = s.lower().strip(" .,;()")
    s = re.sub(r"\bдля\b|\bстанк[аов]\b|\bмод\.?\b|\bсерии\b|\bузла\b","",s)
    return re.sub(r"\s+"," ",s).strip()

# ══════════════════════════════════════════════════════════════════
# 1. ПАРСИНГ Directus-офферы/*.md  — главный источник описаний
# ══════════════════════════════════════════════════════════════════
def _extract_block(text: str, label: str) -> str:
    """Извлекает содержимое блока ```...``` или строки после **Label:**"""
    # Сначала ищем блок в тройных кавычках после label
    pat = re.compile(
        rf"\*\*{re.escape(label)}[^*]*\*\*\s*```(.*?)```",
        re.DOTALL | re.IGNORECASE
    )
    m = pat.search(text)
    if m: return m.group(1).strip()
    # Иначе — строка после label до следующего **
    pat2 = re.compile(
        rf"\*\*{re.escape(label)}[^*]*\*\*[:\s]*([^\n*`]+)",
        re.IGNORECASE
    )
    m2 = pat2.search(text)
    return m2.group(1).strip() if m2 else ""

def parse_directus_offers(folder: Path) -> dict:
    """
    Возвращает словарь {norm_name: enrichment_dict}
    enrichment_dict содержит: desc, meta_title, meta_description,
                               compat, image, specs{}
    """
    result = {}
    if not folder.exists():
        print(f"  [SKIP] папка не найдена: {folder}"); return result

    files = list(folder.glob("*_Directus.md"))
    print(f"  Directus-офферы: {len(files)} файлов")

    for fp in files:
        try:
            text = fp.read_text(encoding="utf-8", errors="ignore")
        except: continue

        name        = _extract_block(text, "Name")
        if not name:
            # fallback: из заголовка файла
            m = re.search(r"## Товар:\s*(.+)", text)
            name = m.group(1).strip() if m else ""
        if not name: continue

        desc        = _extract_block(text, "Description (расширенное)")
        if not desc: desc = _extract_block(text, "Description")
        meta_title  = _extract_block(text, "Meta Title")
        meta_desc   = _extract_block(text, "Meta Description")
        compat      = _extract_block(text, "Compatible Products")
        image       = _extract_block(text, "Image")

        # Тех.характеристики из описания
        specs = {}
        if desc:
            for rx, key in [(TEETH_RE,"Число зубьев"),(MODULE_RE,"Модуль"),
                            (MATERIAL_RE,"Материал"),(HEAT_RE,"Термообработка"),
                            (HRC_RE,"Твёрдость HRC")]:
                m = rx.search(desc)
                if m: specs[key] = m.group(1).strip()

        # Артикул из имени
        sku = extract_sku(name) or extract_sku(desc)

        key = norm(name)
        result[key] = {
            "name": name, "desc": clean(desc),
            "meta_title": meta_title, "meta_desc": meta_desc,
            "compat": compat, "image": image,
            "sku": sku, "specs": specs,
        }
    print(f"    Загружено карточек: {len(result)}")
    return result

# ══════════════════════════════════════════════════════════════════
# 2. ПАРСИНГ DIRECTUS_TKP_549_FULL.csv (Tilda)
# ══════════════════════════════════════════════════════════════════
def parse_tilda_csv(fp: Path) -> dict:
    """→ {norm_name: {price, photo, compat, sku, section, group}}"""
    if not fp.exists(): print(f"  [SKIP] {fp.name}"); return {}
    df = pd.read_csv(fp, dtype=str, low_memory=False).fillna("")
    cols = {c.strip(): c for c in df.columns}
    def col(*names):
        for n in names:
            if n in cols: return cols[n]
            for k in cols:
                if n.lower() in k.lower(): return cols[k]
        return None
    ct = col("Title"); cs = col("SKU"); cp = col("Photo","Фото")
    cpr = col("Price","Цена"); ccat = col("Category","Категория")
    ctabs = col("Tabs:1"); ctext = col("Text")

    result = {}
    for _, row in df.iterrows():
        title = str(row[ct]).strip() if ct else ""
        if not title or title.lower() in ("nan","title"): continue
        price = str(row[cpr]).strip() if cpr else ""
        if price in ("nan","0",""): price = ""
        photo = str(row[cp]).strip() if cp else ""
        if photo in ("nan",""): photo = ""
        sku = str(row[cs]).strip() if cs else ""
        if sku == "nan": sku = ""
        # Compat из Tabs:1
        compat = ""
        if ctabs:
            m = re.search(r"Совместимость:\s*([^\n]+)", str(row[ctabs]))
            if m: compat = m.group(1).replace(";",", ").strip()
        # Категория
        cat_raw = str(row[ccat]).strip() if ccat else ""
        if ">>>" in cat_raw:
            parts = cat_raw.split(">>>")
            section, group = parts[0].strip(), parts[-1].strip()
        else:
            section, group = "", ""
        # Краткое описание
        desc = clean(str(row[ctext])) if ctext else ""

        result[norm(title)] = {
            "price": price, "photo": photo, "compat": compat,
            "sku": sku, "section": section, "group": group,
            "desc": desc,
        }
    print(f"  DIRECTUS_TKP_549_FULL.csv: {len(result)} записей")
    return result

# ══════════════════════════════════════════════════════════════════
# 3. ПАРСИНГ КИНЕМАТИКИ (XLS) — число зубьев, модуль
# ══════════════════════════════════════════════════════════════════
def parse_kinematics_xls(folder: Path) -> dict:
    """
    Ищет все .xls/.xlsx в папке Kinematika-Chertegi рекурсивно.
    Возвращает {артикул_нормализованный: {specs{}}}
    """
    result = {}
    if not folder.exists(): return result
    files = list(folder.rglob("*.xls")) + list(folder.rglob("*.xlsx"))
    print(f"  Кинематика: найдено {len(files)} XLS-файлов")
    loaded = 0
    for fp in files:
        try:
            xl = pd.ExcelFile(fp, engine="xlrd" if str(fp).endswith(".xls") else "openpyxl")
            for sh in xl.sheet_names:
                try:
                    df = xl.parse(sh, dtype=str, header=None).fillna("")
                    # Ищем строки с артикулом (паттерн типа 1М63.02.262)
                    for _, row in df.iterrows():
                        row_str = " ".join(str(v) for v in row.values)
                        sku_match = SKU_RE.search(row_str)
                        if not sku_match: continue
                        sku = sku_match.group(1)
                        specs = {}
                        # Ищем числовые колонки рядом — число зубьев
                        tm = re.search(r"\bz\s*[=:]\s*(\d+)|\bZ\s*=\s*(\d+)|зубьев\s+(\d+)", row_str, re.I)
                        if tm:
                            specs["Число зубьев"] = next(g for g in tm.groups() if g)
                        mm = re.search(r"\bm\s*[=:]\s*([\d,\.]+)|\bмодуль\s*[=:]\s*([\d,\.]+)", row_str, re.I)
                        if mm:
                            specs["Модуль"] = next(g for g in mm.groups() if g)
                        if specs:
                            result[norm(sku)] = specs
                            loaded += 1
                except: pass
        except: pass
    print(f"    Извлечено спецификаций: {loaded}")
    return result

# ══════════════════════════════════════════════════════════════════
# 4. ПАРСИНГ MD-ФАЙЛОВ (имена)
# ══════════════════════════════════════════════════════════════════
_JUNK = re.compile(
    r"^(\d[\d\s\u2009]*[₽/]|\d+\s*(шт|кг|п\.м|комплект)"
    r"|[-–—]{3,}|#{1,6}\s|\||>\s|\*\*[^*]+\*\*:?\s*$)", re.I
)
def parse_md(fp: Path) -> list:
    if not fp.exists(): print(f"  [SKIP] {fp.name}"); return []
    items, cur_name, cur_desc = [], None, []
    for raw in fp.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or _JUNK.match(line): continue
        if len(line) >= 150 and cur_name:
            cur_desc.append(clean(line)); continue
        if re.match(r"^[А-ЯЁа-яёA-Za-z(«\d]", line) and len(line) > 5:
            if cur_name:
                items.append({"name": cur_name, "desc": " ".join(cur_desc)})
            cur_name = line; cur_desc = []
    if cur_name:
        items.append({"name": cur_name, "desc": " ".join(cur_desc)})
    print(f"  {fp.name}: {len(items)} позиций")
    return items

def load_exclude(fp: Path) -> set:
    if not fp.exists(): return set()
    items = parse_md(fp)
    ex = {norm(i["name"]) for i in items}
    print(f"  Исключений: {len(ex)}")
    return ex

# ══════════════════════════════════════════════════════════════════
# 5. ДЕДУПЛИКАЦИЯ
# ══════════════════════════════════════════════════════════════════
def deduplicate(sources: list, exclude: set) -> list:
    seen, result = set(exclude), []
    for src in sources:
        for item in src:
            k = norm(item["name"])
            if not k or k in seen: continue
            seen.add(k); result.append(item)
    return result

# ══════════════════════════════════════════════════════════════════
# 6. ГЕНЕРАЦИЯ META
# ══════════════════════════════════════════════════════════════════
def gen_meta_title(name: str, sku: str, machines: str) -> str:
    """60-90 символов: Название [Артикул] для [модель] | ТД РУССтанкоСбыт"""
    brand = " | ТД РУССтанкоСбыт"
    base  = name
    if sku and sku not in name: base = f"{name} {sku}"
    if machines:
        first_model = machines.split(",")[0].strip()
        if first_model not in base:
            candidate = f"{base} для {first_model}{brand}"
            if len(candidate) <= 90: return candidate
    candidate = f"{base}{brand}"
    if len(candidate) > 90:
        candidate = f"{name[:70].rsplit(' ',1)[0]}{brand}"
    return candidate

def gen_meta_desc(name: str, sku: str, machines: str, specs: dict) -> str:
    """150-200 символов"""
    parts = []
    if sku and sku not in name:
        parts.append(f"Арт. {sku}.")
    if machines:
        parts.append(f"Для станков {machines[:60]}.")
    if specs.get("Число зубьев"):
        parts.append(f"z={specs['Число зубьев']}.")
    if specs.get("Модуль"):
        parts.append(f"m={specs['Модуль']}.")
    if specs.get("Материал"):
        mat = specs["Материал"][:40]
        parts.append(f"Материал: {mat}.")
    parts.append("Купить с доставкой по России. Гарантия. ТД РУССтанкоСбыт.")
    desc = f"{name[:60]}. " + " ".join(parts)
    return desc[:200]

def make_phrases(name: str, machines: str, sku: str) -> str:
    p = []
    if name[:50]: p.append(name[:50])
    words = name.split()
    buy = f"купить {' '.join(words[:3])}"[:50]
    if buy not in p: p.append(buy)
    for m in (machines.split(", ")[:3] if machines else []):
        ph = f"запчасти {m}"[:50]
        if ph not in p: p.append(ph)
    if sku and len(sku) > 3:
        ph = f"{sku[:40]} купить"[:50]
        if ph not in p: p.append(ph)
    for g in ["запчасти токарный станок","купить запчасти с доставкой",
              "комплектующие токарного станка"]:
        if len(p) < 10 and g not in p: p.append(g)
    return "\n".join(p[:10])

# ══════════════════════════════════════════════════════════════════
# 7. СБОРКА СТРОКИ (57 полей)
# ══════════════════════════════════════════════════════════════════
def build_row(item: dict, d_offers: dict, tilda: dict, kine: dict) -> list:
    name  = item["name"].strip()
    key   = norm(name)

    # Ищем обогащение по ключу (точное, потом частичное)
    enriched = d_offers.get(key, {})
    if not enriched:
        # Частичное совпадение — ищем по SKU или подстроке
        for k, v in d_offers.items():
            if key in k or k in key:
                enriched = v; break

    tilda_data = tilda.get(key, {})
    if not tilda_data:
        for k, v in tilda.items():
            if key in k or k in key:
                tilda_data = v; break

    # Собираем поля с приоритетом: Directus-оффер > Tilda > item > auto
    desc  = enriched.get("desc","") or tilda_data.get("desc","") or item.get("desc","")
    sku   = enriched.get("sku","") or tilda_data.get("sku","") or item.get("sku","") or extract_sku(name)
    photo = enriched.get("image","") or tilda_data.get("photo","")
    price = tilda_data.get("price","") or item.get("price","")
    compat = enriched.get("compat","") or tilda_data.get("compat","")
    specs  = enriched.get("specs",{})
    meta_title = enriched.get("meta_title","")
    meta_desc  = enriched.get("meta_desc","")

    # Кинематика — дообогащаем specs если нашли по артикулу
    if sku:
        kspecs = kine.get(norm(sku),{})
        for k,v in kspecs.items():
            if k not in specs: specs[k] = v

    # Машины
    machines = compat or extract_machines(f"{name} {desc}")

    # Категория
    section = enriched.get("section","") or tilda_data.get("section","")
    group   = enriched.get("group","") or tilda_data.get("group","")
    if not section or not group:
        section, group = detect_category(name)

    # Описание — если нет готового, генерим
    if not desc:
        mach_str = f" Применяется на станках {machines}." if machines else ""
        sku_str  = f" Артикул: {sku}." if sku else ""
        desc = (f"{name} — запасная часть для токарного станка.{mach_str}"
                f"{sku_str} Купить с доставкой по России. "
                f"Подбор по каталогу и чертежам заказчика. ТД РУССтанкоСбыт.")

    # META
    if not meta_title:  meta_title = gen_meta_title(name, sku, machines)
    if not meta_desc:   meta_desc  = gen_meta_desc(name, sku, machines, specs)

    phrases = make_phrases(name, machines, sku)

    # Характеристики (до 10 блоков по 3 поля)
    chars = []
    if machines:  chars.append(("Совместимые станки",  machines, ""))
    if sku:       chars.append(("Артикул",             sku,      ""))
    for spec_name, spec_val in list(specs.items())[:7]:
        chars.append((spec_name, spec_val, ""))
    while len(chars) < 10: chars.append(("","",""))

    row = [
        name, desc, "шт.", price, "руб.",
        photo, section, group, "под заказ", sku,
        "ТД РУССтанкоСбыт", phrases,
        "","","",
        "","","","","","","","","","",  # 10 оптовых цен
    ]
    for n_c, v_c, u_c in chars[:10]:
        row += [n_c, v_c, u_c]
    row += [meta_title, meta_desc]  # col 56-57

    assert len(row) == 57, f"Длина строки: {len(row)}"
    return row

# ══════════════════════════════════════════════════════════════════
# 8. ЗАПИСЬ XLSX
# ══════════════════════════════════════════════════════════════════
def write_xlsx(items: list, d_offers: dict, tilda: dict, kine: dict, out: Path):
    wb = Workbook(); ws = wb.active; ws.title = "Лист1"
    ws.append(PROMPORTAL_HEADERS)
    hfill = PatternFill("solid", start_color="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    rfont = Font(name="Arial", size=10)
    enriched_count = 0
    for item in items:
        row = build_row(item, d_offers, tilda, kine)
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = rfont
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        # Считаем обогащённые (есть нормальное описание)
        if len(row[1]) > 100: enriched_count += 1

    for col, w in {"A":65,"B":100,"F":50,"G":38,"H":45,"J":30,"L":45,"BA":60,"BB":80}.items():
        try: ws.column_dimensions[col].width = w
        except: pass

    ws2 = wb.create_sheet("Товарные группы")
    ws2.append(["Ид","Наименование","Ид родителя"])
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return enriched_count

# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════
def main():
    print("="*65)
    print("  ТД РУССтанкоСбыт — Сборка каталога promportal.su v3")
    print("="*65)

    print("\n[1/6] Исключения...")
    exclude = load_exclude(EXCLUDE_FILE)

    print("\n[2/6] Directus-офферы (описания + meta + specs)...")
    d_offers = parse_directus_offers(DOFFERS_DIR)

    print("\n[3/6] Tilda CSV (цены + фото + совместимость)...")
    tilda = parse_tilda_csv(TILDA_CSV)

    print("\n[4/6] Кинематика XLS...")
    kine = parse_kinematics_xls(KINEMATICS_DIR)

    print("\n[5/6] MD-файлы номенклатуры...")
    md_sources = [
        parse_md(NOMENCLATURA),
        parse_md(RZN_OFFERS),
        parse_md(PARTS_1M63),
    ]
    # Добавляем позиции из Directus-офферов как источник имён
    doffer_items = [{"name": v["name"], "desc": v["desc"],
                     "sku": v["sku"], "price": ""} for v in d_offers.values()]

    all_sources = [doffer_items] + md_sources
    merged = deduplicate(all_sources, exclude)
    print(f"\n  Итого уникальных новых позиций: {len(merged)}")

    print("\n[6/6] Генерирую XLSX...")
    enriched = write_xlsx(merged, d_offers, tilda, kine, OUTPUT_FILE)

    print(f"\n✅ Готово!")
    print(f"   Файл:       {OUTPUT_FILE}")
    print(f"   Строк:      {len(merged)}")
    print(f"   С описанием (>100 симв): {enriched}")
    print(f"   Без описания (авто):     {len(merged)-enriched}")

    # Статистика категорий
    cats = {}
    for item in merged:
        _, g = detect_category(item["name"])
        cats[g] = cats.get(g,0)+1
    print("\n📊 Топ-15 товарных групп:")
    for g, cnt in sorted(cats.items(), key=lambda x:-x[1])[:15]:
        print(f"   {cnt:4d}  {g}")

if __name__ == "__main__":
    main()
