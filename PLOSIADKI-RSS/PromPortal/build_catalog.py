"""
build_catalog.py — Сборка каталога для promportal.su
ТД РУССтанкоСбыт

Запуск:
    pip install openpyxl pandas
    python build_catalog.py

Выход: promportal_upload_ДАТА.xlsx
"""

import re
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

# ─── ПУТИ К ФАЙЛАМ ────────────────────────────────────────────────────────────
BASE = r"C:\GitHub-Repositories\Katalog-RSS"

SOURCES = {
    "nomenclatura":  os.path.join(BASE, "NOMENCLATURA RSS-2026.md"),
    "rzn_offers":    os.path.join(BASE, "PLOSIADKI-RSS", "PromPortal", "RZN-список офферов.md"),
    "catalog_xlsx":  os.path.join(BASE, "CATALOG_OLEG_FINAL_100.xlsx"),
    "parts_1m63":    os.path.join(BASE, "RZN-сайт",
                         "Запасные части для токарно-винторезных станков мод. 1М63Н(1М63),16К40 (1А64),1Н65(1М65),РТ117,РТ817.md"),
}
EXCLUDE_FILE   = os.path.join(BASE, "PLOSIADKI-RSS", "PromPortal", "СПИСОК ОФФЕРОВ.md")
TEMPLATE_FILE  = os.path.join(BASE, "PLOSIADKI-RSS", "PromPortal", "PromPortal-ШАБЛОН.xlsx")
OUTPUT_FILE    = f"promportal_upload_{date.today().strftime('%Y%m%d')}.xlsx"

# ─── МОДЕЛИ СТАНКОВ (для извлечения из текста) ────────────────────────────────
MACHINE_PATTERNS = [
    r"1[МмMm][63]{2}[НнBбBFФ\d]*",
    r"1[НнNn][65]{2}[^\s,]*",
    r"1[МмMm][65]{2}[^\s,]*",
    r"16[КкKk]20[^\s,]*",
    r"16[КкKk]30[^\s,]*",
    r"16[КкKk]40[^\s,]*",
    r"16[МмMm]30[^\s,]*",
    r"16[РрRr]25[^\s,]*",
    r"1[КкKk]62[^\s,]*",
    r"РТ\d{3}[^\s,]*",
    r"ДИП[35]\d{2}[^\s,]*",
    r"165[^\s,]*",
    r"163[^\s,]*",
    r"117[^\s,]*",
    r"817[^\s,]*",
    r"2М\d{2,3}[^\s,]*",
    r"6[РрTт]\d{2}[^\s,]*",
    r"6[НнNn]\d{2}[^\s,]*",
    r"1П\d{3}[^\s,]*",
    r"СА\d{3}[^\s,]*",
]
MACHINE_RE = re.compile("|".join(MACHINE_PATTERNS), re.IGNORECASE)

# ─── КАТЕГОРИИ (по ключевым словам) ──────────────────────────────────────────
CATEGORY_MAP = [
    (["шпиндел", "шпиндл"],                         "Шпиндельная бабка"),
    (["фартук"],                                     "Фартук"),
    (["каретк", "суппорт"],                          "Суппорт и каретка"),
    (["задняя бабка", "задней бабк", "бабка задн"],  "Задняя бабка"),
    (["шестерн", "зубчат", "колесо", "венец", "шлицев"],  "Шестерни и зубчатые колёса"),
    (["вал ", "валик", "вал-"],                      "Валы"),
    (["гайк", "маточн"],                             "Гайки"),
    (["винт"],                                       "Винты и ходовые винты"),
    (["люнет"],                                      "Люнеты"),
    (["патрон"],                                     "Патроны"),
    (["резцедержат"],                                "Резцедержатели"),
    (["пиноль"],                                     "Задняя бабка"),
    (["муфт"],                                       "Муфты"),
    (["ремонт", "капитальн"],                        "Ремонт и восстановление"),
    (["кулачок", "кулачки"],                         "Кулачки"),
    (["лимб"],                                       "Лимбы и нониусы"),
    (["клин"],                                       "Клинья"),
    (["копир"],                                      "Копиры"),
    (["маслоуказ", "смазк"],                         "Системы смазки"),
]

def detect_category(name: str) -> str:
    name_l = name.lower()
    for keywords, cat in CATEGORY_MAP:
        if any(kw in name_l for kw in keywords):
            return cat
    return "Запасные части для станков"

def extract_machines(text: str) -> str:
    found = MACHINE_RE.findall(text)
    unique = list(dict.fromkeys(m.strip(".,;") for m in found))
    return ", ".join(unique[:8])  # не больше 8 моделей

# ─── ПОИСКОВЫЕ ФРАЗЫ (для promportal) ────────────────────────────────────────
def make_search_phrases(name: str, machines: str) -> list[str]:
    """До 10 фраз, каждая ≤ 50 символов"""
    phrases = []
    # 1. Первые 50 символов названия
    short = name[:50].strip()
    phrases.append(short)
    # 2. "купить [первое слово(а)]"
    words = name.split()
    buy_phrase = f"купить {' '.join(words[:3])}"[:50]
    phrases.append(buy_phrase)
    # 3. Модели станков
    if machines:
        for m in machines.split(", ")[:4]:
            p = f"запчасти {m}"[:50]
            if p not in phrases:
                phrases.append(p)
    # 4. Общие ключевые фразы
    generic = ["запчасти токарный станок", "купить запчасти с доставкой"]
    for g in generic:
        if len(phrases) < 10 and g not in phrases:
            phrases.append(g)
    return phrases[:10]

# ─── ПАРСИНГ MD-ФАЙЛОВ ───────────────────────────────────────────────────────
PRICE_RE    = re.compile(r"[\d\s\u2009]+[₽руб]")
DIGITS_ONLY = re.compile(r"^\d[\d\s\u2009.,]*[₽]?$")
SHORT_JUNK  = re.compile(r"^(\d+\s+комплект|[\d\s\u2009]+[₽/].*)$")

def is_title_line(line: str) -> bool:
    if len(line) < 5:                   return False
    if DIGITS_ONLY.match(line):         return False
    if SHORT_JUNK.match(line):          return False
    if re.match(r"^\d+\s*(шт|кг|п\.м)", line, re.I): return False
    if line[0].isdigit() and "₽" in line: return False
    # Должна начинаться с буквы (рус/лат)
    return bool(re.match(r"^[А-ЯЁа-яёA-Za-z(]", line))

def parse_md(filepath: str) -> list[dict]:
    if not os.path.exists(filepath):
        print(f"  [SKIP] не найден: {filepath}")
        return []
    items = []
    current_title = None
    current_desc  = []
    with open(filepath, encoding="utf-8") as f:
        lines = [l.rstrip("\r\n") for l in f]
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if len(line) >= 120:
            # Длинная строка — описание к предыдущему товару
            if current_title:
                current_desc.append(line)
            continue
        if is_title_line(line):
            if current_title:
                items.append({"name": current_title, "desc": " ".join(current_desc)})
            current_title = line
            current_desc  = []
        # Иначе — цена/ед. изм. — пропускаем
    if current_title:
        items.append({"name": current_title, "desc": " ".join(current_desc)})
    print(f"  Прочитано из {os.path.basename(filepath)}: {len(items)} позиций")
    return items

def parse_exclude_md(filepath: str) -> set[str]:
    """Читает СПИСОК ОФФЕРОВ.md → set названий для исключения"""
    if not os.path.exists(filepath):
        return set()
    items = parse_md(filepath)
    return {normalize(i["name"]) for i in items}

def normalize(s: str) -> str:
    """Нормализация для дедупликации"""
    return re.sub(r"\s+", " ", s.lower().strip(" .,;()"))

# ─── ПАРСИНГ XLSX-КАТАЛОГА ───────────────────────────────────────────────────
def parse_catalog_xlsx(filepath: str) -> list[dict]:
    if not os.path.exists(filepath):
        print(f"  [SKIP] не найден: {filepath}")
        return []
    items = []
    df = pd.read_excel(filepath, dtype=str).fillna("")
    # Ищем колонки по смыслу
    cols = {c.lower(): c for c in df.columns}
    name_col  = next((cols[c] for c in cols if "наимен" in c or "title" in c or "name" in c), None)
    desc_col  = next((cols[c] for c in cols if "описан" in c or "desc" in c), None)
    sku_col   = next((cols[c] for c in cols if "артикул" in c or "sku" in c or "код" in c), None)
    compat_col = next((cols[c] for c in cols if "совмест" in c or "compat" in c or "станок" in c or "модел" in c), None)
    price_col  = next((cols[c] for c in cols if "цена" in c or "price" in c), None)

    if not name_col:
        print(f"  [WARN] {os.path.basename(filepath)}: колонка с названием не найдена, пропуск")
        return []
    for _, row in df.iterrows():
        name = str(row[name_col]).strip()
        if not name or name.lower() in ("nan", "наименование", "название"):
            continue
        items.append({
            "name":  name,
            "desc":  str(row[desc_col]).strip()   if desc_col   else "",
            "sku":   str(row[sku_col]).strip()    if sku_col    else "",
            "compat": str(row[compat_col]).strip() if compat_col else "",
            "price": str(row[price_col]).strip()  if price_col  else "",
        })
    print(f"  Прочитано из {os.path.basename(filepath)}: {len(items)} позиций")
    return items

# ─── СБОРКА И ДЕДУПЛИКАЦИЯ ───────────────────────────────────────────────────
def build_merged(sources_data: list[list[dict]], exclude_set: set[str]) -> list[dict]:
    seen    = set(exclude_set)  # исключаем уже загруженные
    result  = []
    total_input = sum(len(s) for s in sources_data)
    for source in sources_data:
        for item in source:
            key = normalize(item["name"])
            if key in seen:
                continue
            seen.add(key)
            result.append(item)
    print(f"\n  Всего на входе: {total_input}")
    print(f"  Уже на портале (исключено): {len(exclude_set)}")
    print(f"  Уникальных новых позиций: {len(result)}")
    return result

# ─── ГЕНЕРАЦИЯ XLSX ───────────────────────────────────────────────────────────
HEADERS = [
    "Наименование", "Описание", "Единица измерения", "Цена", "Валюта",
    "Изображение", "Раздел", "Товарная группа", "Наличие", "Код товара",
    "Производитель", "Поисковые фразы", "Цена от", "Цена со скидкой",
    "Срок скидки",
    "Оптовая цена 1", "Оптовый мин заказ 1",
    "Оптовая цена 2", "Оптовый мин заказ 2",
    "Оптовая цена 3", "Оптовый мин заказ 3",
    "Оптовая цена 4", "Оптовый мин заказ 4",
    "Оптовая цена 5", "Оптовый мин заказ 5",
]
# + 10 блоков характеристик
for i in range(1, 11):
    HEADERS += [f"Название характеристики {i}", f"Значение характеристики {i}",
                f"Измерение характеристики {i}"]

def build_row(item: dict) -> list:
    name    = item.get("name", "")
    desc    = item.get("desc", "")
    sku     = item.get("sku", "")
    compat  = item.get("compat", "")
    price   = item.get("price", "")

    # Извлекаем модели станков из всех текстовых полей
    machines = extract_machines(f"{name} {desc} {compat}")
    category = detect_category(name)
    phrases  = make_search_phrases(name, machines)
    phrases_str = "\n".join(phrases)  # promportal принимает разделитель \n

    # Генерация описания если нет
    if not desc and machines:
        desc = (f"Запасные части для токарных станков {machines}. "
                f"Купить {name.lower()} с доставкой по России. "
                f"Оригинальные комплектующие. Гарантия качества. "
                f"ТД РУССтанкоСбыт.")

    row = [
        name,           # Наименование
        desc,           # Описание
        "шт.",          # Единица измерения
        price,          # Цена
        "руб.",         # Валюта
        "",             # Изображение
        "Запасные части для станков",  # Раздел
        category,       # Товарная группа
        "под заказ",    # Наличие
        sku,            # Код товара
        "ТД РУССтанкоСбыт",  # Производитель
        phrases_str,    # Поисковые фразы
        "",             # Цена от
        "",             # Цена со скидкой
        "",             # Срок скидки
    ]
    # 5 оптовых цен — пусто
    for _ in range(10):
        row.append("")

    # Характеристика 1: Совместимые станки
    if machines:
        row += ["Совместимые станки", machines, ""]
    else:
        row += ["", "", ""]
    # Остальные 9 характеристик — пусто
    for _ in range(9):
        row += ["", "", ""]

    return row

def write_xlsx(items: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    # Заголовок
    ws.append(HEADERS)
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Данные
    row_font = Font(name="Arial", size=10)
    for item in items:
        row = build_row(item)
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = row_font
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    # Ширина колонок
    col_widths = {"A": 60, "B": 80, "G": 30, "H": 35, "L": 40}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Лист "Товарные группы" (пустой, нужен для валидации)
    ws2 = wb.create_sheet("Товарные группы")
    ws2.append(["Ид", "Наименование", "Ид родителя"])

    wb.save(output_path)
    print(f"\n✅ Файл сохранён: {output_path}")
    print(f"   Строк данных: {len(items)}")

# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("Сборка каталога для promportal.su — ТД РУССтанкоСбыт")
    print("=" * 60)

    # 1. Читаем список исключений
    print("\n[1/4] Читаю уже загруженные позиции (исключения)...")
    exclude = parse_exclude_md(EXCLUDE_FILE)
    print(f"  Исключений: {len(exclude)}")

    # 2. Читаем все источники
    print("\n[2/4] Читаю источники...")
    all_sources = []

    # MD-файлы
    for key in ["nomenclatura", "rzn_offers", "parts_1m63"]:
        path = SOURCES[key]
        print(f"  → {os.path.basename(path)}")
        all_sources.append(parse_md(path))

    # XLSX-каталог
    print(f"  → {os.path.basename(SOURCES['catalog_xlsx'])}")
    all_sources.append(parse_catalog_xlsx(SOURCES["catalog_xlsx"]))

    # 3. Мержим и дедублируем
    print("\n[3/4] Дедупликация и фильтрация...")
    merged = build_merged(all_sources, exclude)

    # 4. Генерируем XLSX
    print(f"\n[4/4] Генерирую {OUTPUT_FILE}...")
    write_xlsx(merged, OUTPUT_FILE)

    print("\nГотово! Загрузи файл на promportal.su через 'Импорт товаров'.")

if __name__ == "__main__":
    main()
