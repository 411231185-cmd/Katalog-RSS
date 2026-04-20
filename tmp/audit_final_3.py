#!/usr/bin/env python3
"""Check overlap of problem categories and get full examples"""

import openpyxl
import re

FINAL = "PLOSIADKI-RSS/PromPortal/Каталоги/PromPortal-FINAL-with-descriptions.xlsx"
wb = openpyxl.load_workbook(FINAL, read_only=True, data_only=True)
ws = wb.active

SERVICES_MARKER = "Подобрать конкретную запчасть"

# Classify each row
classes = {}
for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    name = str(row[2] or '')
    desc = str(row[4] or '')
    if not desc.strip():
        classes[row_num] = 'empty'
        continue

    tags = set()
    if desc.strip().startswith('ООО') or desc.strip().startswith('ТД РУС'):
        tags.add('marketing')
    if 'Смотрите также:' in desc:
        tags.add('smotrite')
    if SERVICES_MARKER in desc:
        tags.add('has_services')
    if 'ТЕХНИЧЕСКИЕ ХАРАКТЕРИСТИКИ' in desc or 'Назначение\n' in desc or 'Назначение\r' in desc:
        tags.add('directus_format')
    if 'Позиция' in desc and 'кинематической схеме' in desc:
        tags.add('kinematic')
    # Check for gen_desc_5 format (our approved generator output)
    if 'применяется в' in desc.lower() and 'Поставляется' in desc:
        tags.add('gen_desc_format')

    classes[row_num] = tags

# Count combinations
from collections import Counter
combos = Counter()
for row_num, tags in classes.items():
    if isinstance(tags, str):
        combos[tags] += 1
    else:
        key = '+'.join(sorted(tags))
        combos[key] += 1

print("=== DESCRIPTION TYPE COMBINATIONS ===")
for combo, cnt in combos.most_common(30):
    print(f"  {cnt:4d}  {combo}")

# Get a full example of "smotrite" type
print("\n=== FULL EXAMPLE: 'smotrite' type (Row 122) ===")
wb2 = openpyxl.load_workbook(FINAL, read_only=True, data_only=True)
ws2 = wb2.active
for row_num, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
    if row_num == 122:
        print(f"Name: {row[2]}")
        print(f"Desc:\n{row[4]}")
        break
wb2.close()

# Count how many "smotrite" ALSO have services block
smot_with_svc = 0
smot_without_svc = 0
for row_num, tags in classes.items():
    if isinstance(tags, set) and 'smotrite' in tags:
        if 'has_services' in tags:
            smot_with_svc += 1
        else:
            smot_without_svc += 1

print(f"\n=== 'Смотрите также' rows: with services={smot_with_svc}, without={smot_without_svc} ===")

# Count descriptions that look like our gen_desc_5 output
gen_count = sum(1 for tags in classes.values() if isinstance(tags, set) and 'gen_desc_format' in tags)
print(f"gen_desc_format (our template): {gen_count}")

# Kinematic data descriptions
kin_count = sum(1 for tags in classes.values() if isinstance(tags, set) and 'kinematic' in tags)
print(f"kinematic position data: {kin_count}")

wb.close()
