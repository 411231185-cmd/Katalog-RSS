#!/usr/bin/env python3
"""Audit description_new quality in PromPortal-FINAL-with-descriptions.xlsx"""

import openpyxl
import re
import json
from collections import Counter

FINAL = "PLOSIADKI-RSS/PromPortal/Каталоги/PromPortal-FINAL-with-descriptions.xlsx"

wb = openpyxl.load_workbook(FINAL, read_only=True, data_only=True)
ws = wb.active

# Read header
header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"=== HEADER ({len(header)} cols) ===")
for i, h in enumerate(header[:10]):
    print(f"  col{i}: {h}")

# Check if source_flag exists
sf_idx = None
for i, h in enumerate(header):
    if h and 'source_flag' in str(h).lower():
        sf_idx = i
        print(f"\n  source_flag found at col{i}")
        break
if sf_idx is None:
    print("\n  *** source_flag column NOT FOUND ***")

# Find description_new column
dn_idx = None
for i, h in enumerate(header):
    if h and 'description_new' in str(h).lower():
        dn_idx = i
        break
print(f"  description_new at col{dn_idx}")

# Find name column
name_idx = 2  # Наименование

SERVICES_MARKER = "Подобрать конкретную запчасть"
CAPS_WORDS = ["ПАТРОНОВ", "ПОДШИПНИКОВ", "ЦЕНТРОВ", "СУППОРТОВ", "ПЛАНШАЙБ",
              "ШВП", "ВИНТОВ", "ВАЛОВ", "ВТУЛОК", "ШЕСТЕРЁН", "ЛЮНЕТОВ",
              "ЗАЩИТНЫХ КОЖУХОВ", "КАБИНЕТНЫХ ЗАЩИТ", "ВКЛАДЫШЕЙ", "ЗАХВАТОВ"]

stats = {
    'total': 0,
    'has_desc': 0,
    'empty_desc': 0,
    'has_services_block': 0,
    'missing_services_block': 0,
    'has_html_tags': 0,
    'starts_with_ooo': 0,  # old marketing copy
    'has_hyperlinks': 0,
    'caps_not_linked': 0,
    'short_desc': 0,  # < 50 chars before services block
    'very_long': 0,   # > 3000 chars
}

categories = Counter()
problems = []
samples_by_type = {}

for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    stats['total'] += 1
    name = str(row[name_idx] or '')
    desc = str(row[dn_idx] or '') if dn_idx is not None and row[dn_idx] else ''

    if not desc.strip():
        stats['empty_desc'] += 1
        problems.append((row_num, name[:60], 'EMPTY description_new'))
        continue

    stats['has_desc'] += 1

    # Check for services block
    if SERVICES_MARKER in desc:
        stats['has_services_block'] += 1
    else:
        stats['missing_services_block'] += 1
        if stats['missing_services_block'] <= 10:
            problems.append((row_num, name[:60], 'MISSING services block'))

    # Check for HTML tags
    if re.search(r'<[a-zA-Z/][^>]*>', desc):
        stats['has_html_tags'] += 1
        if stats['has_html_tags'] <= 5:
            problems.append((row_num, name[:60], f'HAS HTML: {desc[:100]}'))

    # Check starts with ООО or marketing copy
    if desc.strip().startswith('ООО') or desc.strip().startswith('ТД РУС'):
        stats['starts_with_ooo'] += 1
        if stats['starts_with_ooo'] <= 5:
            problems.append((row_num, name[:60], f'MARKETING COPY: {desc[:120]}'))

    # Check for hyperlinks in services block
    if SERVICES_MARKER in desc:
        services_part = desc[desc.index(SERVICES_MARKER):]
        if '<a ' in services_part or 'href=' in services_part:
            stats['has_hyperlinks'] += 1
        else:
            stats['caps_not_linked'] += 1

    # Check description length (before services block)
    if SERVICES_MARKER in desc:
        tech_part = desc[:desc.index(SERVICES_MARKER)].strip()
    else:
        tech_part = desc.strip()

    if len(tech_part) < 50:
        stats['short_desc'] += 1
        if stats['short_desc'] <= 5:
            problems.append((row_num, name[:60], f'SHORT tech part ({len(tech_part)} chars): {tech_part[:80]}'))

    if len(desc) > 3000:
        stats['very_long'] += 1

    # Categorize by part type for sampling
    name_lower = name.lower()
    ptype = 'other'
    for kw in ['станок', 'вал', 'шестерн', 'колесо', 'винт', 'гайк', 'муфт', 'втулк',
               'патрон', 'ролик', 'головк', 'швп', 'шарико', 'насос', 'кулачк', 'диск',
               'блок', 'суппорт', 'каретк', 'фартук', 'коробк', 'шпиндел', 'люнет',
               'центр', 'планшайб', 'венец', 'пружин', 'вилк', 'клин', 'шкив']:
        if kw in name_lower:
            ptype = kw
            break
    categories[ptype] += 1

    # Save 2 samples per type
    if ptype not in samples_by_type:
        samples_by_type[ptype] = []
    if len(samples_by_type[ptype]) < 2:
        samples_by_type[ptype].append((row_num, name[:80], tech_part[:200] if tech_part else ''))

wb.close()

print(f"\n=== STATS (total {stats['total']} rows) ===")
for k, v in stats.items():
    print(f"  {k}: {v}")

print(f"\n=== CATEGORIES ===")
for cat, cnt in categories.most_common(20):
    print(f"  {cat}: {cnt}")

print(f"\n=== PROBLEMS (first examples) ===")
for row_num, name, issue in problems[:30]:
    print(f"  Row {row_num}: [{name}] -> {issue}")

print(f"\n=== SAMPLES BY TYPE (tech part only) ===")
for ptype in sorted(samples_by_type.keys()):
    print(f"\n--- {ptype} ---")
    for row_num, name, tech in samples_by_type[ptype]:
        print(f"  Row {row_num}: {name}")
        print(f"    Tech: {tech[:200]}")
